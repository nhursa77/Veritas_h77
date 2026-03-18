from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    print("ERROR: Nedostaje paket openpyxl. Instaliraj: pip install openpyxl")
    raise SystemExit(1) from exc


DEFAULT_INPUT = Path(
    "izvori/dokazno/eu/curia/vjm_iate/iate_popis_svih_jezika/izvor.xlsx"
)
DEFAULT_RAW_OUTPUT = Path(
    "izvori/operativno/eu/curia/vjm_iate/iate_popis_svih_jezika_raw.json"
)
DEFAULT_STRUCTURE_OUTPUT = Path(
    "izvori/operativno/eu/curia/vjm_iate/iate_popis_svih_jezika_struktura.json"
)


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def _read_sheet(ws: Any) -> tuple[dict[str, Any], dict[str, Any]]:
    first_row = next(
        ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column, values_only=True),
        (),
    )
    columns = ["" if cell is None else str(cell) for cell in first_row]

    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(
        ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, values_only=True),
        start=2,
    ):
        values = [_json_safe(cell) for cell in row]
        rows.append(
            {
                "row_index": row_index,
                "values": values,
            }
        )

    row_count = len(rows)
    column_count = len(columns)

    raw_sheet = {
        "worksheet": ws.title,
        "row_count": row_count,
        "column_count": column_count,
        "columns": columns,
        "rows": rows,
    }
    structure_sheet = {
        "worksheet": ws.title,
        "row_count": row_count,
        "column_count": column_count,
        "columns": columns,
    }
    return raw_sheet, structure_sheet


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")


def build_export(input_path: Path, raw_output: Path, structure_output: Path) -> tuple[int, int]:
    workbook = openpyxl.load_workbook(filename=input_path, read_only=True, data_only=True)

    raw_sheets: list[dict[str, Any]] = []
    structure_sheets: list[dict[str, Any]] = []
    total_rows = 0

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        raw_sheet, structure_sheet = _read_sheet(ws)
        raw_sheets.append(raw_sheet)
        structure_sheets.append(structure_sheet)
        total_rows += structure_sheet["row_count"]

    raw_payload = {
        "izvor": str(input_path.as_posix()),
        "worksheet_count": len(raw_sheets),
        "total_rows": total_rows,
        "worksheets": raw_sheets,
    }
    structure_payload = {
        "izvor": str(input_path.as_posix()),
        "worksheet_count": len(structure_sheets),
        "total_rows": total_rows,
        "worksheets": structure_sheets,
    }

    _write_json(raw_output, raw_payload)
    _write_json(structure_output, structure_payload)
    workbook.close()
    return len(raw_sheets), total_rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pretvara CURIA XLSX dokazni izvor u sirovi JSON izvoz."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--raw-output", type=Path, default=DEFAULT_RAW_OUTPUT)
    parser.add_argument("--structure-output", type=Path, default=DEFAULT_STRUCTURE_OUTPUT)
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    input_path = args.input
    raw_output = args.raw_output
    structure_output = args.structure_output

    if not input_path.exists():
        print(f"ERROR: Ulazna datoteka ne postoji: {input_path}")
        return 2

    worksheet_count, total_rows = build_export(
        input_path=input_path,
        raw_output=raw_output,
        structure_output=structure_output,
    )

    print(f"WORKSHEETS_COUNT={worksheet_count}")
    print(f"TOTAL_EXPORTED_ROWS={total_rows}")
    print(f"RAW_OUTPUT_PATH={raw_output.as_posix()}")
    print(f"STRUCTURE_OUTPUT_PATH={structure_output.as_posix()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
